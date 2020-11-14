#!/usr/bin/python
# coding:utf-8

"""
@author: NookVoice
@software: PyCharm
@file: table_column_to_excel.py
@time: 2020/10/28 20:56
"""
import logging
import os
import re
import time
import tkinter as tk
from tkinter import ttk, filedialog
from tkinter import simpledialog

import cx_Oracle
import openpyxl as openpyxl
import psycopg2
import pymysql as pymysql


class DIALOG_FILE(tk.Toplevel):
    def __init__(self):
        super().__init__()
        self.title("警告!")
        self.geometry('200x100')
        tk.Label(self, text="文件已存在是否覆盖?").grid(row=1)
        tk.Button(self, text="是", command=self.yes).grid(row=2, column=0)
        tk.Button(self, text="否", command=self.no).grid(row=2, column=1)

    def yes(self):
        self.is_cover = True
        self.destroy()

    def no(self):
        self.is_cover = False
        self.destroy()


class DB_INFO():
    def __init__(self, db_type, db_host, db_port, db_name, db_user, db_passwd):
        self.db_type = db_type
        self.db_host = db_host
        self.db_port = db_port
        self.db_name = db_name
        self.db_user = db_user
        self.db_passwd = db_passwd


class Excel_INFO():
    def __init__(self, file_path, file_name, file_type):
        self.file_path = file_path
        self.file_name = file_name
        self.file_type = file_type
        self.file_name_full = file_path + '/' + file_name + file_type


class TABLE_INFO():
    def __init__(self, table_owner, table_schema, table_name, table_comment):
        self.table_owner = table_owner
        self.table_schema = table_schema
        self.table_name = table_name
        self.table_comment = table_comment

        self.table_name_full = table_schema + '.' + table_name
        self.sheet_name = self.table_name_full
        self.table_link = '=HYPERLINK("#{sheet_name}!A1","{table_name}")'.format(sheet_name=self.sheet_name,
                                                                                      table_name=self.table_name)
    def update_table_link(self):
        self.table_link = '=HYPERLINK("#{sheet_name}!A1","{table_name}")'.format(sheet_name=self.sheet_name,
                                                                                      table_name=self.table_name)

    def update_sheet_name(self, sheet_name):
        self.sheet_name = sheet_name
        # 更新表名对应超链接
        self.table_link()


class BASE_DESK(tk.Tk):

    def __init__(self):
        super().__init__()

    def __init__(self, title):
        super().__init__()
        self.title(title)

        # 初始化变量
        self.init_variable()

        # 初始化数据库类型下拉框
        self.init_cmb()

        # 初始化输入框
        self.init_entry()

    def init_variable(self):
        self.db_types = ["PostgreSQL", "MySQL", "Oracle"]
        self.db_type = ''
        self.db_host = tk.StringVar()
        self.db_port = tk.IntVar()
        self.db_database = tk.StringVar()
        self.db_user = tk.StringVar()
        self.db_passwd = tk.StringVar()

        self.file_path = tk.StringVar()
        self.file_name = tk.StringVar()
        self.file_type = ".xlsx"

    def init_cmb(self):
        # 数据库类型

        tk.Label(self, text="数据库类型: ").grid(row=0, column=0)
        self.cmb = ttk.Combobox(self, values=self.db_types)
        self.cmb.grid(row=0, column=1)
        self.cmb.current(0)
        # 选择后触发操作
        self.cmb.bind("<<ComboboxSelected>>", self.chose_db_type)

    def init_entry(self):
        # 数据库host
        tk.Label(self, text="数据库地址: ").grid(row=1, column=0)
        self.et_host = tk.Entry(self, textvariable=self.db_host)
        self.et_host.grid(row=1, column=1)
        # 输入后触发操作
        self.et_host.bind("<<KeyRelease>>", func=self.update_file_name)

        # 数据库port
        tk.Label(self, text="数据库端口: ").grid(row=2, column=0)
        self.et_port = tk.Entry(self, textvariable=self.db_port)
        self.et_port.grid(row=2, column=1)
        # 输入后触发操作
        self.et_port.bind("<<KeyRelease>>", func=self.update_file_name)

        # 数据库名称
        tk.Label(self, text="数据库名称: ").grid(row=3, column=0)
        self.et_database = tk.Entry(self, textvariable=self.db_database)
        self.et_database.grid(row=3, column=1)
        # 输入后触发操作
        self.et_database.bind("<<KeyRelease>>", func=self.update_file_name)

        # 数据库用户
        tk.Label(self, text="数据库用户: ").grid(row=4, column=0)
        self.et_user = tk.Entry(self, textvariable=self.db_user)
        self.et_user.grid(row=4, column=1)
        # 输入后触发操作
        self.et_user.bind("<<KeyRelease>>", func=self.update_file_name)

        # 数据库密码
        tk.Label(self, text="数据库密码: ").grid(row=5, column=0)
        self.et_passwd = tk.Entry(self, textvariable=self.db_passwd, show='*')
        self.et_passwd.grid(row=5, column=1)
        # 输入后触发操作
        self.et_passwd.bind("<<KeyRelease>>", func=self.update_file_name)

        # 导出文件夹
        tk.Label(self, text="导出文件夹: ").grid(row=6, column=0)
        self.et_file_path = tk.Entry(self, textvariable=self.file_path)
        self.et_file_path.grid(row=6, column=1)
        tk.Button(self, text="浏览", command=self.select_path).grid(row=6, column=2)

        # 导出文件名

        tk.Label(self, text="导出文件名: ").grid(row=7, column=0)
        self.et_file_path = tk.Entry(self, textvariable=self.file_name)
        self.et_file_path.grid(row=7, column=1)
        tk.Label(self, text=self.file_type).grid(row=7, column=2)

        # 提交与退出
        tk.Button(self, text="提交", command=self.submit).grid(row=8, column=0)
        tk.Button(self, text="退出", command=self.quit).grid(row=8, column=1)

    def chose_db_type(self, root):
        func_map = {
            "PostgreSQL": self.default_postgres,
            "MySQL": self.default_mysql,
            "Oracle": self.default_oracle
        }

        func = func_map.get(self.cmb['values'][self.cmb.current()])
        func()

    def default_postgres(self):
        self.db_type = "postgres"
        self.db_host.set("localhost")
        self.db_port.set(5432)
        self.db_database.set("postgres")
        self.db_user.set("postgres")
        self.db_passwd.set("")
        self.file_path.set(os.path.join(os.path.expanduser("~"), "Desktop"))
        self.file_name.set("表结构-" + self.db_host.get() + "-" + str(self.db_port.get()) + "-"
                           + self.db_database.get() + "-" + self.db_user.get()
                           + time.strftime('%Y%m%d', time.localtime(time.time())))

    def default_mysql(self):
        self.db_type = "mysql"
        self.db_host.set("localhost")
        self.db_port.set(3306)
        self.db_database.set("mysql")
        self.db_user.set("root")
        self.db_passwd.set("")
        self.file_path.set(os.path.join(os.path.expanduser("~"), "Desktop"))
        self.file_name.set("表结构-" + self.db_host.get() + "-" + str(self.db_port.get()) + "-"
                           + self.db_database.get() + "-" + self.db_user.get()
                           + time.strftime('%Y%m%d', time.localtime(time.time())))

    def default_oracle(self):
        self.db_type = "oracle"
        self.db_host.set("localhost")
        self.db_port.set(1521)
        self.db_database.set("orcl")
        self.db_user.set("sysdba")
        self.db_passwd.set("")
        self.file_path.set(os.path.join(os.path.expanduser("~"), "Desktop"))
        self.file_name.set("表结构-" + self.db_host.get() + "-" + str(self.db_port.get()) + "-"
                           + self.db_database.get() + "-" + self.db_user.get()
                           + time.strftime('%Y%m%d', time.localtime(time.time())))

    def update_file_name(self, root):
        self.file_name.set("表结构-" + self.db_host.get() + "-" + str(self.db_port.get()) + "-"
                           + self.db_database.get() + "-" + self.db_user.get()
                           + time.strftime('%Y%m%d', time.localtime(time.time())))

    def select_path(self):
        self.file_path.set(filedialog.askdirectory())

    def touch_file(self, file_info: Excel_INFO):
        if os.access(file_info.file_name_full, os.F_OK):
            dialog_file = DIALOG_FILE()
            self.wait_window(dialog_file)
            if not dialog_file.is_cover:
                return False

        if file_info.file_type == '.xlsx':
            target_file = openpyxl.Workbook()
            target_file.save(file_info.file_name_full)
            return True

    def submit(self):
        db_info = DB_INFO(db_type=self.db_type,
                          db_host=self.db_host.get(), db_port=self.db_port.get(), db_name=self.db_database.get(),
                          db_user=self.db_user.get(), db_passwd=self.db_passwd.get())
        excel_info = Excel_INFO(self.file_path.get(), self.file_name.get(), self.file_type)

        # 检查并创建文件
        self.touch_file(excel_info)

        export_to_file(db_info, excel_info)


def get_db_cur(db_info):
    db_type = db_info.db_type
    if db_type == 'postgres':
        conn = psycopg2.connect(host=db_info.db_host, port=db_info.db_port, database=db_info.db_name,
                                user=db_info.db_user, password=db_info.db_passwd)
        cur = conn.cursor()
        return cur

    elif db_type == 'mysql':
        conn = pymysql.connect(host=db_info.db_host, port=db_info.db_port, database=db_info.db_name,
                               user=db_info.db_user, password=db_info.db_passwd)
        cur = conn.cursor()
        return cur
    elif db_type == 'oracle':
        # cx_Oracle.connect("hr", "welcome", "localhost/orclpdb1")
        conn = cx_Oracle.connect(db_info.db_user, db_info.db_passwd,
                                 db_info.db_host + ":" + str(db_info.db_port) + '/' + db_info.db_name)
        cur = conn.cursor()
        return cur
    else:
        print("数据库类型错误,请重新选择!")


def get_table_list_postgres(cur):
    sql_text = '''
select 
    lower(u.usename) as table_owner
    ,lower(n.nspname) as table_schema
    ,lower(c.relname) as table_name
    ,cast(obj_description(c.oid,'pg_class') as varchar) as table_comment
from pg_class c 
join pg_namespace n on c.relnamespace = n.oid 
left join pg_user u on c.relowner = u.usesysid 
where lower(n.nspname) not in ('pg_toast','information_schema','pg_catalog')
order by n.nspname ,c.relname
'''
    cur.execute(sql_text)
    table_list = cur.fetchall()

    return table_list


def get_table_list_mysql(cur):
    sql_text = '''
    select 
table_schema as table_owner 
,table_schema as table_schema
,table_name  as table_name
,table_comment as table_comment 
from information_schema.tables t 
where lower(table_schema) <> lower('information_schema')
  and lower(table_schema) <> lower('mysql')
  and lower(table_schema) <> lower('performance_schema')
  and lower(table_schema) <> lower('sys')
order by t.table_schema,t.table_name 
    '''

    cur.execute(sql_text)
    table_list = cur.fetchall()

    return table_list


def get_table_list_oracle(cur):
    sql_text = '''
    select 
t.owner as table_owner 
,t.owner as table_schema
,t.table_name  as table_name
,t_1.comments as table_comment 
from all_tables t 
left join all_table_comments t_1 on t.table_name = t_1.table_name 
order by t.owner,t.table_name 
    '''

    cur.execute(sql_text)
    table_list = cur.fetchall()

    return table_list


def get_table_list(db_info):
    table_list = []

    # 获取数据库连接
    cur = get_db_cur(db_info)

    if db_info.db_type == 'postgres':
        table_list = get_table_list_postgres(cur)
        return table_list

    elif db_info.db_type == 'mysql':
        table_list = get_table_list_mysql(cur)
        return table_list
    elif db_info.db_type == 'oracle':
        table_list = get_table_list_oracle(cur)
        return table_list
    else:
        print("数据库类型错误,请重新选择!")

    cur.close()
    return table_list


def check_sheet_name(dict_sheet_name, sheet_name):
    # Excelsheet页限制最大31长度,预留三位序号位及一位下划线
    max_sheet_name_len = 31 - 4
    # sheet名称不支持符号
    p = re.compile(r'[:\\/?*\[\]]')
    sheet_name = re.sub(p, "", sheet_name)

    if len(sheet_name) > max_sheet_name_len:
        # 右侧截取最大长度并去除'.'符号
        sheet_name_new = sheet_name[- max_sheet_name_len:].strip('.')

        logging.warning(
            'length of : {sheet_name} out of {max_sheet_name_len} , rename to : {sheet_name_new} '.format(
                sheet_name=sheet_name.ljust(50), max_sheet_name_len=max_sheet_name_len,
                sheet_name_new=sheet_name_new))

        sheet_name = sheet_name_new

    # 判断名称是否已存在
    if sheet_name in dict_sheet_name.keys():
        dict_sheet_name[sheet_name] += 1
        sheet_name_new = sheet_name + '_' + str(dict_sheet_name[sheet_name]).zfill(3)
        logging.warning(
            'name of : {sheet_name} already exists, rename to : {sheet_name_new} '.format(
                sheet_name=sheet_name.ljust(50),
                sheet_name_new=sheet_name_new))
        sheet_name = sheet_name_new
    else:
        dict_sheet_name.setdefault(sheet_name, 1)

    return sheet_name


def get_column_list_postgres(cur, table_info):
    sql_text = '''
    
    '''
    pass


def get_column_list_mysql(cur, table_info):
    pass


def get_column_list_oracle(cur, table_info):
    pass


def get_column_list(db_info, table_info):
    column_list = []

    # 获取数据库连接
    cur = get_db_cur(db_info)

    if db_info.db_type == 'postgres':
        column_list = get_column_list_postgres(cur,table_info)
        return column_list

    elif db_info.db_type == 'mysql':
        column_list = get_column_list_mysql(cur,table_info)
        return column_list
    elif db_info.db_type == 'oracle':
        column_list = get_column_list_oracle(cur,table_info)
        return column_list
    else:
        print("数据库类型错误,请重新选择!")

    cur.close()
    return column_list


def add_sheet_table(db_info, table_info, target_file, index=0):
    sheet_table = target_file.create_sheet(index=index, title=table_info.sheet_name)

    link_catalog = '=HYPERLINK("#{sheet_name}!A1","返回首页")'.format(sheet_name="目录")

    # 获取表字段
    column_list = get_column_list(db_info,table_info)


def add_excel_sheet(target_file, db_info):
    # 获取表清单
    table_list = get_table_list(db_info)

    # 目录页
    sheet_catalog = target_file.create_sheet(index=0, title="目录")

    sheet_catalog.cell(row=1, column=1, value='序号')
    sheet_catalog.cell(row=1, column=2, value='表所有者')
    sheet_catalog.cell(row=1, column=3, value='模式名')
    sheet_catalog.cell(row=1, column=4, value='表名')
    sheet_catalog.cell(row=1, column=5, value='表中文名')

    # 遍历插入
    for i, table in enumerate(table_list):
        table_info = TABLE_INFO(table_owner=table[0], table_schema=table[1], table_name=table[2],
                                table_comment=table[3])

        sheet_name = table_info.table_name_full

        # 表sheet页名称字典,并记录该名称重复次数
        dict_sheet_name = dict()
        sheet_name = check_sheet_name(dict_sheet_name, sheet_name)

        # 更新表对应sheet名
        table_info.update_sheet_name(sheet_name)

        sheet_catalog.cell(row=i + 2, column=1, value=i + 1)
        sheet_catalog.cell(row=i + 2, column=2, value=table_info.table_owner)
        sheet_catalog.cell(row=i + 2, column=3, value=table_info.table_schema)
        sheet_catalog.cell(row=i + 2, column=4, value=table_info.table_link)
        sheet_catalog.cell(row=i + 2, column=5, value=table_info.table_comment)

        # 增加每张表的sheet页
        add_sheet_table(target_file=target_file, db_info=db_info, table_info=table_info, index=i + 1)


def export_to_file(db_info, excel_info):
    # 目标文件
    target_file = openpyxl.load_workbook(excel_info.file_name_full)

    # 根据表清单写入Excel
    add_excel_sheet(target_file, db_info)

    target_file.save(filename=excel_info.file_name_full)
    target_file.close()


def main():
    base_desk = BASE_DESK(title="表字段导出Excel工具")

    base_desk.mainloop()


if __name__ == "__main__":
    # 日志等级
    logging.basicConfig(level=logging.INFO)
    main()
